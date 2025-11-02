#!/usr/bin/env python3
"""
Import script for Lake Charles customers from CSV to Arctic Ice Solutions
Transforms CSV data to match the Customer model and creates Excel file for bulk import
"""

import csv
import uuid
import re
import json
from pathlib import Path
try:
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def parse_address(address_str):
    """Parse address string into components: street, city, state, zip"""
    
    parts = re.split(r'\s{2,}', address_str.strip())
    
    if len(parts) >= 3:
        street = parts[0]
        city = parts[1]
        state_zip = parts[2]
        
        state_zip_match = re.match(r'([A-Z]{2})\s+(\d{5})', state_zip)
        if state_zip_match:
            state = state_zip_match.group(1)
            zip_code = state_zip_match.group(2)
        else:
            state = "LA"
            zip_code = "70601"
    else:
        street = address_str
        city = "Lake Charles"
        state = "LA"
        zip_code = "70601"
    
    return street, city, state, zip_code

def generate_email(customer_name):
    """Generate email from customer name"""
    clean_name = re.sub(r'[^a-zA-Z0-9\s]', '', customer_name.lower())
    clean_name = re.sub(r'\s+', '', clean_name)
    return f"contact@{clean_name}example.com"

def transform_csv_to_excel():
    """Transform the Lake Charles CSV to Excel format for Arctic Ice import"""
    
    csv_path = "/home/ubuntu/attachments/91c94e95-0026-402c-b534-5445b7359cff/real_lake_charles_customers_90.csv"
    
    transformed_data = []
    
    with open(csv_path, 'r', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        
        for row in reader:
            if not row.get('Customer') or not row['Customer'].strip():
                continue
                
            customer_name = row['Customer'].strip()
            address = row['Address'].strip()
            latitude = float(row['Latitude'])
            longitude = float(row['Longitude'])
            
            street, city, state, zip_code = parse_address(address)
            
            customer_id = str(uuid.uuid4())
            
            customer_record = {
                'id': customer_id,
                'name': customer_name,
                'contact_person': 'Manager',
                'phone': '(337) 555-0100',
                'email': generate_email(customer_name),
                'address': street,
                'city': city,
                'state': state,
                'zip_code': zip_code,
                'location_id': 'loc_2',  # Lake Charles location
                'credit_limit': 10000.0,
                'payment_terms': 30,
                'is_active': True,
                'coordinates': f'{{"lat": {latitude}, "lng": {longitude}}}'
            }
            
            transformed_data.append(customer_record)
    
    print(f"Processing {len(transformed_data)} customers from CSV")
    print(f"openpyxl available: {HAS_OPENPYXL}")
    
    if HAS_OPENPYXL:
        output_path = "/home/ubuntu/repos/arctic-ice-solutions/backend/lake_charles_customers_import.xlsx"
        wb = Workbook()
        ws = wb.active
        
        headers = list(transformed_data[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for row_idx, record in enumerate(transformed_data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=record[header])
        
        wb.save(output_path)
        print(f"Created Excel file: {output_path}")
    else:
        output_path = "/home/ubuntu/repos/arctic-ice-solutions/backend/lake_charles_customers_import.csv"
        with open(output_path, 'w', newline='', encoding='utf-8') as file:
            if transformed_data:
                writer = csv.DictWriter(file, fieldnames=transformed_data[0].keys())
                writer.writeheader()
                writer.writerows(transformed_data)
        print(f"Created CSV file: {output_path}")
    
    print(f"Transformed {len(transformed_data)} customers")
    
    print("\nSample of transformed data:")
    for i, record in enumerate(transformed_data[:3]):
        print(f"\nCustomer {i+1}:")
        print(f"  Name: {record['name']}")
        print(f"  Address: {record['address']}, {record['city']}, {record['state']} {record['zip_code']}")
        print(f"  Location ID: {record['location_id']}")
        print(f"  Coordinates: {record['coordinates']}")
    
    return output_path, len(transformed_data)

if __name__ == "__main__":
    try:
        output_file, count = transform_csv_to_excel()
        print(f"\nSuccess! Created {output_file} with {count} customers ready for import.")
    except Exception as e:
        print(f"Error: {e}")
        raise
